# backend/app/services/excel_import_service.py

from __future__ import annotations

from io import BytesIO
import re
import unicodedata
from typing import Any, Dict, List, Tuple, Optional, Set

from openpyxl import load_workbook

from app.core.database import get_db_connection


# -------------------------
# Normalization / cleaning
# -------------------------

def _clean_header(x: Any) -> str:
    s = "" if x is None else str(x)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_name(name: str) -> str:
    """
    Normalize names to ASCII-only characters, removing accents and special chars.
    Handles Turkish chars explicitly.
    """
    if name is None:
        return ""
    if not isinstance(name, str):
        name = str(name)

    name = name.strip()
    if not name:
        return ""

    replacements = {
        "ı": "i", "İ": "I",
        "ğ": "g", "Ğ": "G",
        "ş": "s", "Ş": "S",
        "ç": "c", "Ç": "C",
        "ö": "o", "Ö": "O",
        "ü": "u", "Ü": "U",
    }
    for turkish_char, ascii_char in replacements.items():
        name = name.replace(turkish_char, ascii_char)

    normalized = unicodedata.normalize("NFD", name)
    ascii_name = "".join(
        ch for ch in normalized
        if unicodedata.category(ch) != "Mn"
    )

    return ascii_name.strip()


def _norm_key(name: str) -> str:
    """
    Key used for caching/matching (case-insensitive, whitespace-normalized)
    after ASCII normalization.
    """
    s = _normalize_name(name or "")
    s = re.sub(r"\s+", " ", s).strip()
    return s.casefold()


def _split_people(cell: Any) -> List[str]:
    """
    Split a cell containing one or more names:
      "X1, X2" => ["X1","X2"]
    Handles: comma, semicolon, slash, ampersand, "and", newlines.
    Applies name normalization.
    """
    if cell is None:
        return []
    s = str(cell).strip()
    if not s:
        return []

    parts = re.split(r",|;|/|\band\b|&|\n", s, flags=re.IGNORECASE)
    out: List[str] = []
    for p in parts:
        p = _normalize_name(p.strip())
        if p:
            out.append(p)
    return out


def _split_course_code(cell: Any) -> str:
    """
    Normalize course codes like "COMP 100" -> "COMP100"
    Keep slashes like "COMP423/523"
    """
    if cell is None:
        return ""
    s = str(cell).strip()
    if not s:
        return ""
    s = re.sub(r"\s+", "", s)
    return s.upper()


def _to_str(x: Any) -> Optional[str]:
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None


def _to_int(x: Any, default: int = 0) -> int:
    """
    Extract first numeric token (supports "12", "12.0", "12 students") -> int
    """
    if x is None:
        return default
    try:
        if isinstance(x, str):
            s = x.strip()
            if not s:
                return default
            m = re.findall(r"[\d.]+", s)
            if not m:
                return default
            x = m[0]
        return int(float(x))
    except Exception:
        return default


def _sheet_to_dicts(wb, sheet_name: str) -> List[Dict[str, Any]]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_clean_header(h) for h in rows[0]]
    out: List[Dict[str, Any]] = []

    for r in rows[1:]:
        if r is None:
            continue
        row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row.values()):
            continue
        out.append(row)

    return out


def _get_first(row: Dict[str, Any], *keys: str) -> Any:
    """
    Return the first existing key from row.
    Useful because Excel headers sometimes include trailing spaces.
    """
    for k in keys:
        if k in row:
            return row.get(k)
    return None


# -------------------------
# Main import
# -------------------------

def import_comp_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    Imports from the COMP Excel workbook (idempotent-ish):
      - TA Needs Planning:
          * course (upsert-ish)
          * course_professor (DELETE+INSERT to mirror Excel; supports "X1, X2" -> 2 professors)
          * preferred TAs -> course_preferred_ta + professor_preferred_ta
          * assigned TAs (names) -> ta_assignment
      - COMP TA List:
          * ta (insert/update)
          * thesis advisor(s) -> professor + ta_thesis_advisor + ta_preferred_professor

    Returns ONLY:
      summary: {
        tas_inserted, tas_updated,
        professors_inserted, professors_updated,
        courses_inserted, courses_updated,
        preferences_updated
      }
      changes: { new_tas, updated_tas, new_professors, new_courses, updated_courses, notes }
    """

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    planning_sheet = "TA Needs Planning"
    ta_list_sheet = "COMP TA List"
    if planning_sheet not in wb.sheetnames or ta_list_sheet not in wb.sheetnames:
        raise ValueError(f"Workbook must contain sheets: '{planning_sheet}' and '{ta_list_sheet}'")

    planning_rows = _sheet_to_dicts(wb, planning_sheet)
    ta_rows = _sheet_to_dicts(wb, ta_list_sheet)

    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)

    # -----------------------
    # small helpers (local)
    # -----------------------
    def load_cache(table: str, id_col: str, name_col: str, key_fn) -> Dict[str, int]:
        cur.execute(f"SELECT {id_col} AS idv, {name_col} AS namev FROM {table}")
        cache: Dict[str, int] = {}
        for r in (cur.fetchall() or []):
            key = key_fn(r["namev"])
            if not key:
                continue
            rid = int(r["idv"])
            if key not in cache or rid < cache[key]:
                cache[key] = rid
        return cache

    def get_or_create_by_name(
        table: str,
        id_col: str,
        name_col: str,
        raw_value: str,
        cache: Dict[str, int],
        normalize_fn,
        key_fn,
    ) -> Tuple[int, bool]:
        nm = normalize_fn(raw_value)
        key = key_fn(nm)
        if not key:
            raise ValueError(f"{table}: empty name")

        if key in cache:
            return cache[key], False

        # schema may not enforce UNIQUE => safe lookup first
        cur.execute(
            f"SELECT {id_col} AS idv FROM {table} WHERE {name_col}=%s ORDER BY {id_col} ASC LIMIT 1",
            (nm,),
        )
        row = cur.fetchone()
        if row:
            rid = int(row["idv"])
            cache[key] = rid
            return rid, False

        cur.execute(f"INSERT INTO {table} ({name_col}) VALUES (%s)", (nm,))
        rid = int(cur.lastrowid)
        cache[key] = rid
        return rid, True

    # -----------------------
    # load caches
    # -----------------------
    try:
        prof_cache = load_cache("professor", "professor_id", "name", _norm_key)
        ta_cache = load_cache("ta", "ta_id", "name", _norm_key)

        # course cache uses course_code normalization (COMP 100 -> COMP100)
        cur.execute("SELECT course_id, course_code FROM course")
        course_cache: Dict[str, int] = {}
        for r in (cur.fetchall() or []):
            code = _split_course_code(r["course_code"])
            if not code:
                continue
            cid = int(r["course_id"])
            if code not in course_cache or cid < course_cache[code]:
                course_cache[code] = cid

        # -----------------------
        # track changes for UI
        # -----------------------
        new_tas: set[str] = set()
        updated_tas: set[str] = set()
        new_professors: set[str] = set()
        new_courses: set[str] = set()
        updated_courses: set[str] = set()

        skipped_preferred_tokens: set[str] = set()
        skipped_assigned_tokens: set[str] = set()
        notes: List[str] = []

        tas_inserted = tas_updated = 0
        courses_inserted = courses_updated = 0
        professors_inserted = 0
        professors_updated = 0  # we don't update professor rows in this importer
        preferences_updated = 0  # combined count for preferences/assignments tables

        # -----------------------
        # 1) COMP TA List
        # -----------------------
        for row in ta_rows:
            raw_name = _to_str(row.get("NAME"))
            if not raw_name:
                continue
            name = _normalize_name(raw_name)
            tkey = _norm_key(name)
            if not tkey:
                continue

            payload = (
                name,
                _to_str(row.get("PROGRAM")),
                "PhD"
                if ((_to_str(row.get("MS/PhD")) or "MS").strip().casefold() in ["phd", "ph.d", "ph.d."])
                else "MS",
                _to_str(row.get("BACKGROUND")),
                _to_str(row.get("ADMIT TERM")),
                _to_int(row.get("STANDING"), default=0),
                _to_str(row.get("NOTES")),
                _to_str(row.get("BS SCHOOL/PROGRAM")),
                _to_str(row.get("MS SCHOOL/PROGRAM")),
            )

            if tkey in ta_cache:
                ta_id = ta_cache[tkey]
                cur.execute(
                    """
                    UPDATE ta
                    SET name=%s, program=%s, level=%s, background=%s, admit_term=%s,
                        standing=%s, notes=%s, bs_school_program=%s, ms_school_program=%s
                    WHERE ta_id=%s
                    """,
                    (*payload, ta_id),
                )
                tas_updated += 1
                updated_tas.add(name)
            else:
                cur.execute(
                    """
                    INSERT INTO ta
                      (name, program, level, background, admit_term, standing, notes, bs_school_program, ms_school_program)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    payload,
                )
                ta_id = int(cur.lastrowid)
                ta_cache[tkey] = ta_id
                tas_inserted += 1
                new_tas.add(name)

            thesis_advisor_cell = _to_str(row.get("THESIS ADVISOR"))
            if thesis_advisor_cell:
                for adv in _split_people(thesis_advisor_cell):
                    prof_id, created = get_or_create_by_name(
                        "professor", "professor_id", "name",
                        adv, prof_cache, _normalize_name, _norm_key
                    )
                    if created:
                        professors_inserted += 1
                        new_professors.add(_normalize_name(adv))

                    cur.execute(
                        "INSERT IGNORE INTO ta_thesis_advisor (ta_id, professor_id) VALUES (%s, %s)",
                        (ta_id, prof_id),
                    )
                    preferences_updated += cur.rowcount

                    cur.execute(
                        "INSERT IGNORE INTO ta_preferred_professor (ta_id, professor_id) VALUES (%s, %s)",
                        (ta_id, prof_id),
                    )
                    preferences_updated += cur.rowcount

        # -----------------------
        # 2) TA Needs Planning
        # -----------------------
        for row in planning_rows:
            course_code = _split_course_code(_get_first(row, "Course"))
            if not course_code:
                continue

            # get/create course
            if course_code in course_cache:
                course_id = course_cache[course_code]
                courses_updated += 1
                updated_courses.add(course_code)
            else:
                # schema may not enforce UNIQUE => safe lookup first
                cur.execute(
                    "SELECT course_id FROM course WHERE course_code=%s ORDER BY course_id ASC LIMIT 1",
                    (course_code,),
                )
                found = cur.fetchone()
                if found:
                    course_id = int(found["course_id"])
                    course_cache[course_code] = course_id
                    courses_updated += 1
                    updated_courses.add(course_code)
                else:
                    cur.execute("INSERT INTO course (course_code) VALUES (%s)", (course_code,))
                    course_id = int(cur.lastrowid)
                    course_cache[course_code] = course_id
                    courses_inserted += 1
                    new_courses.add(course_code)

            # update course fields
            cur.execute(
                """
                UPDATE course
                SET ps_lab_sections=%s,
                    enrollment_capacity=%s,
                    actual_enrollment=%s,
                    num_tas_requested=%s,
                    assigned_tas_count=%s
                WHERE course_id=%s
                """,
                (
                    _to_str(_get_first(row, "PS/Lab Sections")),
                    _to_int(_get_first(row, "Enrollment Capacity"), default=0),
                    _to_int(_get_first(row, "Actual Enrollment"), default=0),
                    _to_int(_get_first(row, "Number of TAs requested for Spring 2025"), default=0),
                    _to_int(_get_first(row, "Assigned TAs for Spring 2025 (number)"), default=0),
                    course_id,
                ),
            )

            # faculty list (supports "X1, X2")
            faculty_names = _split_people(_get_first(row, "Faculty"))

            # rewrite course_professor exactly as Excel says
            cur.execute("DELETE FROM course_professor WHERE course_id=%s", (course_id,))
            for prof_name in faculty_names:
                prof_id, created = get_or_create_by_name(
                    "professor", "professor_id", "name",
                    prof_name, prof_cache, _normalize_name, _norm_key
                )
                if created:
                    professors_inserted += 1
                    new_professors.add(_normalize_name(prof_name))

                cur.execute(
                    "INSERT INTO course_professor (course_id, professor_id) VALUES (%s, %s)",
                    (course_id, prof_id),
                )

            # preferred TAs (skip requirements text)
            pref_cell = _get_first(row, "Preferred TAs (or requirements)", "Preferred TAs (or requirements) ")
            for token in _split_people(pref_cell):
                if re.fullmatch(r"\+?\d+", token.strip()):
                    continue
                k = _norm_key(token)
                if k not in ta_cache:
                    skipped_preferred_tokens.add(token.strip())
                    continue
                ta_id = ta_cache[k]

                cur.execute(
                    "INSERT IGNORE INTO course_preferred_ta (course_id, ta_id) VALUES (%s, %s)",
                    (course_id, ta_id),
                )
                preferences_updated += cur.rowcount

                for prof_name in faculty_names:
                    if not prof_name:
                        continue
                    prof_id, _ = get_or_create_by_name(
                        "professor", "professor_id", "name",
                        prof_name, prof_cache, _normalize_name, _norm_key
                    )
                    cur.execute(
                        "INSERT IGNORE INTO professor_preferred_ta (professor_id, ta_id) VALUES (%s, %s)",
                        (prof_id, ta_id),
                    )
                    preferences_updated += cur.rowcount

            # assigned TAs (names) -> ta_assignment
            assigned_cell = _get_first(row, "Assigned TAs for Spring 2025 (names)")
            for token in _split_people(assigned_cell):
                k = _norm_key(token)
                if k not in ta_cache:
                    skipped_assigned_tokens.add(token.strip())
                    continue
                ta_id = ta_cache[k]
                cur.execute(
                    "INSERT IGNORE INTO ta_assignment (ta_id, course_id) VALUES (%s, %s)",
                    (ta_id, course_id),
                )
                preferences_updated += cur.rowcount

        # notes
        if skipped_preferred_tokens:
            notes.append(
                f"Skipped {len(skipped_preferred_tokens)} preferred tokens that didn't match any TA name."
            )
        if skipped_assigned_tokens:
            notes.append(
                f"Skipped {len(skipped_assigned_tokens)} assigned TA names that didn't match any TA name."
            )

        conn.commit()

        # keep response light if lists are huge
        def cap(lst: List[str], n: int = 300) -> List[str]:
            return lst if len(lst) <= n else (lst[:n] + [f"... (+{len(lst)-n} more)"])

        return {
            "ok": True,
            "summary": {
                "tas_inserted": tas_inserted,
                "tas_updated": tas_updated,
                "professors_inserted": professors_inserted,
                "professors_updated": professors_updated,
                "courses_inserted": courses_inserted,
                "courses_updated": courses_updated,
                "preferences_updated": preferences_updated,
            },
            "changes": {
                "new_tas": cap(sorted(new_tas)),
                "updated_tas": cap(sorted(updated_tas)),
                "new_professors": cap(sorted(new_professors)),
                "new_courses": cap(sorted(new_courses)),
                "updated_courses": cap(sorted(updated_courses)),
                "notes": notes,
            },
        }

    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
